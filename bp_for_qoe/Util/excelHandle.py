import xlrd
import openpyxl

'''
@Author:xiayifan
@Function:提供excel相关接口功能，包括从excel中获取数据集，向excel写入权值矩阵,从excel读取权值

'''
DataSetPath = './static/1.xlsx'
WidthPath = './static/save.xlsx'
DataSetPathTest = '../static/1.xlsx'
SavePathTest = '../static/save.xlsx'


def GetDataSetAndLabelsFormExcel():
    workbook = xlrd.open_workbook(DataSetPathTest)  # 通过xlrd打开excel文件，1.xlsx目前为测试文件
    sheet = workbook.sheet_by_index(0)  # 取excel文件的第一张表
    row = sheet.nrows  # 记录行列数
    casesRaw = []
    labelsRaw = []
    for i in range(row):  # 取每行元素 制作case 与 label
        caseRaw = []
        labelRaw = []
        delayInEx = sheet.row(i)[0].value
        shakeInEx = sheet.row(i)[1].value
        packetInEx = sheet.row(i)[2].value
        bandwidthInEx = sheet.row(i)[3].value
        labelInEx = sheet.row(i)[4].value
        caseRaw.append(delayInEx)
        caseRaw.append(shakeInEx)
        caseRaw.append(packetInEx)
        caseRaw.append(bandwidthInEx)
        casesRaw.append(caseRaw)
        labelRaw.append(labelInEx)
        labelsRaw.append(labelRaw)
    return [casesRaw, labelsRaw]


def GetNNFromExcel():
    workbook = xlrd.open_workbook(SavePathTest)
    sheet = workbook.sheet_by_index(0)
    # row = sheet.nrows
    # col = sheet.ncols
    mat = [sheet.row_values(0), sheet.row_values(1), sheet.row_values(2), sheet.row_values(3)]
    # for i in range(row):
    #     hang = []
    #     for j in range(col):
    #         hang.append(sheet.row(i)[j])
    #     mat.append(hang)
    return mat


def InsertNNToExcel(width, atrnorm, labnorm):
    workbook = openpyxl.Workbook()
    sheetW = workbook.active
    sheetW.title = 'width'
    k = 0
    for wd in width:
        k = k + 1
        m = 1
        for i in range(len(wd)):
            for j in range(len(wd[i])):
                sheetW.cell(row=k, column=m, value=wd[i][j])
                m = m + 1
    l = 1
    k = 3
    for no in atrnorm:
        for i in no:
            sheetW.cell(row=k, column=l, value=i)
            l += 1
    k = 4
    sheetW.cell(row=k, column=1, value=labnorm[0])
    sheetW.cell(row=k, column=2, value=labnorm[1])
    workbook.save(SavePathTest)
